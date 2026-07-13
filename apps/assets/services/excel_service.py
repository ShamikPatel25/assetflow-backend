import io
import json
from datetime import datetime
from decimal import Decimal

import openpyxl
from django.core.exceptions import ObjectDoesNotExist

from apps.assets.models import Asset, AssetCategory
from django.conf import settings


def export_assets_to_excel(queryset):
    """Exports a queryset of Assets to an Excel workbook."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Assets"

    headers = [
        "Asset Code", "Name", "Category Code", "Brand", "Model", 
        "Serial Number", "Status", "Condition", "Purchase Cost", 
        "Currency", "Purchase Date", "Warranty Expiry Date", "Metadata"
    ]
    ws.append(headers)

    for asset in queryset:
        ws.append([
            asset.asset_code,
            asset.name,
            asset.category.code if asset.category else "",
            asset.brand or "",
            asset.model or "",
            asset.serial_number or "",
            asset.status,
            asset.condition,
            str(asset.purchase_cost) if asset.purchase_cost else "",
            asset.currency,
            asset.purchase_date.strftime("%Y-%m-%d") if asset.purchase_date else "",
            asset.warranty_expiry_date.strftime("%Y-%m-%d") if asset.warranty_expiry_date else "",
            ", ".join([f"{k}: {v}" for k, v in asset.metadata.items()]) if asset.metadata else ""
        ])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def import_assets_from_excel(file_obj, user=None):
    """
    Imports new Assets from an uploaded Excel file.
    Only creates new assets. Returns errors for existing asset_codes.
    """
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.active

    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
    
    errors = []
    success_count = 0
    skipped_count = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_data = dict(zip(headers, row))
        
        # Skip completely empty rows based on Asset Code
        if not row_data.get("Asset Code"):
            continue

        asset_code = str(row_data.get("Asset Code")).strip()
        
        # 1. Check uniqueness (Bulk creation only, skip updates)
        existing_asset = Asset.objects.filter(asset_code=asset_code).first()
        if existing_asset:
            excel_serial = str(row_data.get("Serial Number", "")).strip() if row_data.get("Serial Number") else None
            if existing_asset.serial_number == excel_serial:
                skipped_count += 1
                continue
            else:
                errors.append(f"Row {row_idx}: Asset with code '{asset_code}' already exists with a different serial number.")
                continue

        row_errors = []

        serial_number = str(row_data.get("Serial Number", "")).strip() if row_data.get("Serial Number") else None
        if serial_number and Asset.objects.filter(serial_number=serial_number).exists():
            row_errors.append(f"Asset with serial number '{serial_number}' already exists")

        # 2. Check Category Code
        category_code = row_data.get("Category Code")
        category = None
        if not category_code:
            row_errors.append("Category Code is required")
        else:
            try:
                category = AssetCategory.objects.get(code=str(category_code).strip())
            except ObjectDoesNotExist:
                row_errors.append(f"Category '{category_code}' does not exist")

        # 3. Process fields and create asset
        def parse_date(val):
            if not val:
                return None
            if isinstance(val, datetime):
                return val.date()
            try:
                return datetime.strptime(str(val).strip(), "%Y-%m-%d").date()
            except ValueError:
                raise ValueError(f"Invalid date format '{val}', expected YYYY-MM-DD")

        purchase_date = None
        warranty_date = None
        
        try:
            purchase_date = parse_date(row_data.get("Purchase Date"))
        except ValueError as e:
            row_errors.append(str(e))
            
        try:
            warranty_date = parse_date(row_data.get("Warranty Expiry Date"))
        except ValueError as e:
            row_errors.append(str(e))
            
        if purchase_date and warranty_date and warranty_date < purchase_date:
            row_errors.append("Warranty Expiry Date cannot be earlier than Purchase Date")

        cost_val = row_data.get("Purchase Cost")
        purchase_cost = None
        if cost_val:
            try:
                purchase_cost = Decimal(str(cost_val))
            except Exception:
                row_errors.append(f"Invalid Purchase Cost '{cost_val}'")

        raw_meta = row_data.get("Metadata")
        metadata_val = str(raw_meta).strip() if  raw_meta else ""
        metadata = {}
        if metadata_val:
            try:
                metadata = json.loads(metadata_val)
                if not isinstance(metadata, dict):
                    row_errors.append("Metadata JSON must be a dictionary object")
            except json.JSONDecodeError:
                clean_val = metadata_val.strip('{} ')
                pairs = clean_val.split(',')
                for pair in pairs:
                    if ':' in pair:
                        k, v = pair.split(':', 1)
                        if k.strip() and v.strip():
                            metadata[k.strip()] = v.strip()
                    else:
                        if pair.strip():
                            row_errors.append("Metadata must be formatted as 'Key: Value, Key2: Value2' or valid JSON")
                            break

        if row_errors:
            errors.append(f"Row {row_idx}: " + ", ".join(row_errors) + ".")
            continue
        
        try:

            Asset.objects.create(
                asset_code=asset_code,
                name=str(row_data.get("Name", "")).strip(),
                category=category,
                brand=str(row_data.get("Brand", "")).strip() if row_data.get("Brand") else None,
                model=str(row_data.get("Model", "")).strip() if row_data.get("Model") else None,
                serial_number=str(row_data.get("Serial Number", "")).strip() if row_data.get("Serial Number") else None,
                status=str(row_data.get("Status", Asset.Status.AVAILABLE)).strip() if row_data.get("Status") else Asset.Status.AVAILABLE,
                condition=str(row_data.get("Condition", Asset.Condition.NEW)).strip() if row_data.get("Condition") else Asset.Condition.NEW,
                purchase_cost=purchase_cost,
                currency=str(row_data.get("Currency", settings.DEFAULT_CURRENCY)).strip() if row_data.get("Currency") else settings.DEFAULT_CURRENCY,
                purchase_date=purchase_date,
                warranty_expiry_date=warranty_date,
                metadata=metadata,
                created_by=user
            )
            success_count += 1
            
        except Exception as e:
            errors.append(f"Row {row_idx}: Error parsing data - {str(e)}")

    return {
        "success": success_count,
        "skipped": skipped_count,
        "errors": errors
    }
