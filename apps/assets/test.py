"""
Tests for Assets and AssetCategory modules.

Covers:
- Category hierarchy (parent-child creation)
- Asset CRUD with role-based permissions
- Status transitions (AVAILABLE → ALLOCATED → IN_MAINTENANCE → RETIRED)
- Asset code uniqueness enforcement
- Category type validation
- Employee cannot create/edit/delete assets
- Soft-delete behavior
- Edge cases: missing required fields, invalid category references
"""
import pytest
from rest_framework import status

pytestmark = pytest.mark.django_db


# 1. ASSET CATEGORY MANAGEMENT

class TestAssetCategories:
    """Category CRUD, hierarchy, and permissions."""

    url = "/api/v1/asset-categories/"

    def test_hr_can_list_categories(self, hr_api_client, category):
        """HR_MANAGER can list asset categories."""
        response = hr_api_client.get(self.url)
        assert response.status_code == status.HTTP_200_OK

    def test_employee_can_read_categories(self, employee_api_client, category):
        """Employees can read categories (for request dropdown)."""
        response = employee_api_client.get(self.url)
        assert response.status_code == status.HTTP_200_OK

    def test_employee_cannot_create_category(self, employee_api_client):
        """EMPLOYEE cannot create asset categories."""
        response = employee_api_client.post(self.url, data={
            "name": "Hacked", "code": "HACK", "category_type": "HARDWARE",
        })
        assert response.status_code in [status.HTTP_403_FORBIDDEN]

    def test_hr_can_create_top_level_category(self, hr_api_client):
        """HR creates a top-level category (no parent)."""
        response = hr_api_client.post(self.url, data={
            "name": "Hardware", "code": "HW", "category_type": "HARDWARE",
        })
        assert response.status_code == status.HTTP_201_CREATED
        assert response.data["name"] == "Hardware"

    def test_hr_can_create_child_category(self, hr_api_client, category):
        """HR creates a child category under existing parent."""
        response = hr_api_client.post(self.url, data={
            "name": "Gaming Laptops", "code": "GLAP",
            "category_type": "HARDWARE", "parent": str(category.id),
        })
        assert response.status_code == status.HTTP_201_CREATED

    def test_duplicate_category_code_rejected(self, hr_api_client, category):
        """Duplicate code → 400."""
        response = hr_api_client.post(self.url, data={
            "name": "Duplicate", "code": category.code, "category_type": "HARDWARE",
        })
        assert response.status_code == status.HTTP_400_BAD_REQUEST

    def test_category_with_invalid_type(self, hr_api_client):
        """Invalid category_type → 400."""
        response = hr_api_client.post(self.url, data={
            "name": "Bad Type", "code": "BTYPE", "category_type": "UNKNOWN",
        })
        assert response.status_code == status.HTTP_400_BAD_REQUEST


# 2. ASSET CRUD PERMISSIONS

class TestAssetPermissions:
    """Test who can create, read, update, delete assets."""

    url = "/api/v1/assets/"

    def test_unauthenticated_cannot_access_assets(self, api_client, tenant):
        """No JWT → blocked."""
        response = api_client.get(self.url)
        assert response.status_code in [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN]

    def test_employee_can_list_assets(self, employee_api_client, asset):
        """EMPLOYEE can see asset list (read-only)."""
        response = employee_api_client.get(self.url)
        assert response.status_code == status.HTTP_200_OK

    def test_employee_cannot_create_asset(self, employee_api_client, category):
        """EMPLOYEE cannot add an asset to inventory."""
        response = employee_api_client.post(self.url, data={
            "asset_code": "HACK-001", "name": "Hacked",
            "category": str(category.id),
        })
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_hr_can_create_asset(self, hr_api_client, category):
        """HR_MANAGER can add an asset."""
        response = hr_api_client.post(self.url, data={
            "asset_code": "AST-NEW-001", "name": "Dell Monitor",
            "category": str(category.id), "status": "AVAILABLE",
        })
        assert response.status_code == status.HTTP_201_CREATED

    def test_admin_can_create_asset(self, admin_api_client, category):
        """ORG_ADMIN can add an asset."""
        response = admin_api_client.post(self.url, data={
            "asset_code": "AST-ADM-001", "name": "Admin Asset",
            "category": str(category.id), "status": "AVAILABLE",
        })
        assert response.status_code == status.HTTP_201_CREATED

    def test_employee_cannot_delete_asset(self, employee_api_client, asset):
        """EMPLOYEE cannot soft-delete an asset."""
        url = f"{self.url}{asset.id}/"
        response = employee_api_client.delete(url)
        assert response.status_code == status.HTTP_403_FORBIDDEN


# 3. ASSET DATA VALIDATION

class TestAssetValidation:
    """Serializer-level validation edge cases."""

    url = "/api/v1/assets/"

    def test_asset_code_must_be_unique(self, hr_api_client, asset, category):
        """Duplicate asset_code → 400."""
        response = hr_api_client.post(self.url, data={
            "asset_code": asset.asset_code,
            "name": "Duplicate", "category": str(category.id),
        })
        assert response.status_code == status.HTTP_400_BAD_REQUEST

    def test_asset_requires_category(self, hr_api_client):
        """Missing category → 400."""
        response = hr_api_client.post(self.url, data={
            "asset_code": "NO-CAT-001", "name": "No Category",
        })
        assert response.status_code == status.HTTP_400_BAD_REQUEST

    def test_invalid_status_value_rejected(self, hr_api_client, category):
        """Invalid status string → 400."""
        response = hr_api_client.post(self.url, data={
            "asset_code": "BAD-STATUS", "name": "Bad",
            "category": str(category.id), "status": "FLYING",
        })
        assert response.status_code == status.HTTP_400_BAD_REQUEST

    def test_invalid_condition_value_rejected(self, hr_api_client, category):
        """Invalid condition string → 400."""
        response = hr_api_client.post(self.url, data={
            "asset_code": "BAD-COND", "name": "Bad",
            "category": str(category.id), "condition": "EXCELLENT",
        })
        assert response.status_code == status.HTTP_400_BAD_REQUEST

    def test_asset_code_required(self, hr_api_client, category):
        """Missing asset_code → 400."""
        response = hr_api_client.post(self.url, data={
            "name": "No Code", "category": str(category.id),
        })
        assert response.status_code == status.HTTP_400_BAD_REQUEST


# 4. ASSET STATUS TRANSITIONS

class TestAssetStatusTransitions:
    """Status updates via PUT."""

    url = "/api/v1/assets/"

    def test_update_asset_to_in_maintenance(self, hr_api_client, asset):
        """HR can update asset status to IN_MAINTENANCE."""
        url = f"{self.url}{asset.id}/"
        response = hr_api_client.put(url, data={
            "asset_code": asset.asset_code, "name": asset.name,
            "category": str(asset.category.id), "status": "IN_MAINTENANCE",
        })
        assert response.status_code == status.HTTP_200_OK

    def test_update_asset_to_retired(self, hr_api_client, asset):
        """HR can retire an asset."""
        url = f"{self.url}{asset.id}/"
        response = hr_api_client.put(url, data={
            "asset_code": asset.asset_code, "name": asset.name,
            "category": str(asset.category.id), "status": "RETIRED",
        })
        assert response.status_code == status.HTTP_200_OK

    def test_employee_cannot_change_asset_status(self, employee_api_client, asset):
        """EMPLOYEE cannot change asset status."""
        url = f"{self.url}{asset.id}/"
        response = employee_api_client.put(url, data={
            "asset_code": asset.asset_code, "name": asset.name,
            "category": str(asset.category.id), "status": "RETIRED",
        })
        assert response.status_code == status.HTTP_403_FORBIDDEN


# ---------------------------------------------------------------------------
# Excel service + import/export endpoints
# ---------------------------------------------------------------------------

import io
import json
from datetime import date
from decimal import Decimal

import openpyxl

from apps.assets.models import Asset
from apps.assets.services.excel_service import (
    export_assets_to_excel,
    import_assets_from_excel,
)

EXCEL_HEADERS = [
    "Asset Code", "Name", "Category Code", "Brand", "Model",
    "Serial Number", "Status", "Condition", "Purchase Cost",
    "Currency", "Purchase Date", "Warranty Expiry Date", "Metadata",
]


def _build_xlsx(rows, headers=None):
    """Build an in-memory .xlsx workbook and return raw bytes.

    ``rows`` is a list of dicts keyed by header name; missing keys become blank.
    """
    headers = headers or EXCEL_HEADERS
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _load_ws(excel_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    return wb.active


class TestExportAssetsToExcel:
    """Direct tests for export_assets_to_excel()."""

    def test_headers_written(self, asset_factory, category):
        asset_factory(name="Laptop", category=category)
        data = export_assets_to_excel(Asset.objects.all())
        ws = _load_ws(data)
        header_row = [c.value for c in ws[1]]
        assert header_row == EXCEL_HEADERS

    def test_full_asset_row_values(self, asset_factory, category):
        asset_factory(
            name="Full Asset",
            category=category,
            asset_code="EXP-001",
            brand="Dell",
            model="XPS",
            serial_number="SER-EXP-001",
            status="AVAILABLE",
            condition="NEW",
            purchase_cost=Decimal("1500.50"),
            currency="USD",
            purchase_date=date(2023, 1, 15),
            warranty_expiry_date=date(2025, 1, 15),
            metadata={"ram": "16GB", "color": "black"},
        )
        ws = _load_ws(export_assets_to_excel(Asset.objects.all()))
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) == 1
        row = dict(zip(EXCEL_HEADERS, rows[0]))
        assert row["Asset Code"] == "EXP-001"
        assert row["Name"] == "Full Asset"
        assert row["Category Code"] == category.code
        assert row["Brand"] == "Dell"
        assert row["Model"] == "XPS"
        assert row["Serial Number"] == "SER-EXP-001"
        assert row["Status"] == "AVAILABLE"
        assert row["Condition"] == "NEW"
        assert row["Purchase Cost"] == "1500.50"
        assert row["Currency"] == "USD"
        assert row["Purchase Date"] == "2023-01-15"
        assert row["Warranty Expiry Date"] == "2025-01-15"
        assert "ram: 16GB" in row["Metadata"]
        assert "color: black" in row["Metadata"]

    def test_asset_with_empty_optional_fields(self, asset_factory, category):
        """Optional/null fields export as empty strings, not errors."""
        asset_factory(
            name="Minimal",
            category=category,
            asset_code="EXP-MIN",
            brand=None,
            model=None,
            serial_number=None,
            purchase_cost=None,
            purchase_date=None,
            warranty_expiry_date=None,
            metadata={},
        )
        ws = _load_ws(export_assets_to_excel(Asset.objects.all()))
        row = dict(zip(EXCEL_HEADERS, list(ws.iter_rows(min_row=2, values_only=True))[0]))
        assert row["Brand"] in ["", None]
        assert row["Model"] in ["", None]
        assert row["Serial Number"] in ["", None]
        assert row["Purchase Cost"] in ["", None]
        assert row["Purchase Date"] in ["", None]
        assert row["Warranty Expiry Date"] in ["", None]
        assert row["Metadata"] in ["", None]


class TestImportAssetsFromExcel:
    """Direct tests for import_assets_from_excel() — one branch per test."""

    def test_successful_create(self, category, tenant):
        data = _build_xlsx([{
            "Asset Code": "IMP-001", "Name": "New Laptop",
            "Category Code": category.code, "Brand": "HP", "Model": "G8",
            "Serial Number": "SN-IMP-001", "Status": "AVAILABLE",
            "Condition": "NEW", "Purchase Cost": "999.99", "Currency": "USD",
            "Purchase Date": "2023-05-01", "Warranty Expiry Date": "2025-05-01",
        }])
        result = import_assets_from_excel(io.BytesIO(data))
        assert result["success"] == 1
        assert result["skipped"] == 0
        assert result["errors"] == []
        created = Asset.objects.get(asset_code="IMP-001")
        assert created.name == "New Laptop"
        assert created.category_id == category.id
        assert created.purchase_cost == Decimal("999.99")
        assert created.purchase_date == date(2023, 5, 1)

    def test_create_with_defaults(self, category, tenant):
        """Blank status/condition/currency fall back to model defaults."""
        data = _build_xlsx([{
            "Asset Code": "IMP-DEF", "Name": "Defaults",
            "Category Code": category.code,
        }])
        result = import_assets_from_excel(io.BytesIO(data))
        assert result["success"] == 1
        a = Asset.objects.get(asset_code="IMP-DEF")
        assert a.status == Asset.Status.AVAILABLE
        assert a.condition == Asset.Condition.NEW
        assert a.currency == "INR"

    def test_missing_asset_code_row_skipped(self, category, tenant):
        """Rows without Asset Code are silently skipped (not counted)."""
        data = _build_xlsx([{
            "Asset Code": "", "Name": "No Code",
            "Category Code": category.code,
        }])
        result = import_assets_from_excel(io.BytesIO(data))
        assert result["success"] == 0
        assert result["skipped"] == 0
        assert result["errors"] == []

    def test_existing_asset_same_serial_skipped(self, asset_factory, category, tenant):
        asset_factory(asset_code="IMP-EXIST", category=category, serial_number="SN-SAME")
        data = _build_xlsx([{
            "Asset Code": "IMP-EXIST", "Name": "Dup",
            "Category Code": category.code, "Serial Number": "SN-SAME",
        }])
        result = import_assets_from_excel(io.BytesIO(data))
        assert result["skipped"] == 1
        assert result["success"] == 0
        assert result["errors"] == []

    def test_existing_asset_different_serial_error(self, asset_factory, category, tenant):
        asset_factory(asset_code="IMP-DIFF", category=category, serial_number="SN-A")
        data = _build_xlsx([{
            "Asset Code": "IMP-DIFF", "Name": "Dup",
            "Category Code": category.code, "Serial Number": "SN-B",
        }])
        result = import_assets_from_excel(io.BytesIO(data))
        assert result["success"] == 0
        assert result["skipped"] == 0
        assert len(result["errors"]) == 1
        assert "different serial number" in result["errors"][0]

    def test_duplicate_serial_number_error(self, asset_factory, category, tenant):
        asset_factory(asset_code="OTHER", category=category, serial_number="SN-DUP")
        data = _build_xlsx([{
            "Asset Code": "IMP-NEW", "Name": "New",
            "Category Code": category.code, "Serial Number": "SN-DUP",
        }])
        result = import_assets_from_excel(io.BytesIO(data))
        assert result["success"] == 0
        assert any("already exists" in e for e in result["errors"])

    def test_missing_category_code_error(self, tenant):
        data = _build_xlsx([{
            "Asset Code": "IMP-NOCAT", "Name": "No Cat", "Category Code": "",
        }])
        result = import_assets_from_excel(io.BytesIO(data))
        assert result["success"] == 0
        assert any("Category Code is required" in e for e in result["errors"])

    def test_nonexistent_category_error(self, tenant):
        data = _build_xlsx([{
            "Asset Code": "IMP-BADCAT", "Name": "Bad Cat",
            "Category Code": "DOES-NOT-EXIST",
        }])
        result = import_assets_from_excel(io.BytesIO(data))
        assert result["success"] == 0
        assert any("does not exist" in e for e in result["errors"])

    def test_invalid_date_format_error(self, category, tenant):
        data = _build_xlsx([{
            "Asset Code": "IMP-BADDATE", "Name": "Bad Date",
            "Category Code": category.code, "Purchase Date": "01-05-2023",
        }])
        result = import_assets_from_excel(io.BytesIO(data))
        assert result["success"] == 0
        assert any("Invalid date format" in e for e in result["errors"])

    def test_warranty_before_purchase_error(self, category, tenant):
        data = _build_xlsx([{
            "Asset Code": "IMP-WAR", "Name": "War",
            "Category Code": category.code,
            "Purchase Date": "2023-05-01",
            "Warranty Expiry Date": "2022-05-01",
        }])
        result = import_assets_from_excel(io.BytesIO(data))
        assert result["success"] == 0
        assert any("cannot be earlier than Purchase Date" in e for e in result["errors"])

    def test_invalid_purchase_cost_error(self, category, tenant):
        data = _build_xlsx([{
            "Asset Code": "IMP-COST", "Name": "Cost",
            "Category Code": category.code, "Purchase Cost": "not-a-number",
        }])
        result = import_assets_from_excel(io.BytesIO(data))
        assert result["success"] == 0
        assert any("Invalid Purchase Cost" in e for e in result["errors"])

    def test_metadata_valid_json_dict(self, category, tenant):
        data = _build_xlsx([{
            "Asset Code": "IMP-META1", "Name": "Meta",
            "Category Code": category.code,
            "Metadata": json.dumps({"ram": "16GB", "cpu": "i7"}),
        }])
        result = import_assets_from_excel(io.BytesIO(data))
        assert result["success"] == 1
        a = Asset.objects.get(asset_code="IMP-META1")
        assert a.metadata == {"ram": "16GB", "cpu": "i7"}

    def test_metadata_json_not_dict_error(self, category, tenant):
        data = _build_xlsx([{
            "Asset Code": "IMP-META2", "Name": "Meta",
            "Category Code": category.code, "Metadata": "[1, 2, 3]",
        }])
        result = import_assets_from_excel(io.BytesIO(data))
        assert result["success"] == 0
        assert any("must be a dictionary" in e for e in result["errors"])

    def test_metadata_key_value_pairs(self, category, tenant):
        data = _build_xlsx([{
            "Asset Code": "IMP-META3", "Name": "Meta",
            "Category Code": category.code, "Metadata": "ram: 16GB, color: black",
        }])
        result = import_assets_from_excel(io.BytesIO(data))
        assert result["success"] == 1
        a = Asset.objects.get(asset_code="IMP-META3")
        assert a.metadata == {"ram": "16GB", "color": "black"}

    def test_metadata_malformed_error(self, category, tenant):
        data = _build_xlsx([{
            "Asset Code": "IMP-META4", "Name": "Meta",
            "Category Code": category.code, "Metadata": "just a plain string",
        }])
        result = import_assets_from_excel(io.BytesIO(data))
        assert result["success"] == 0
        assert any("Key: Value" in e for e in result["errors"])

    def test_create_with_user(self, category, org_admin_user, tenant):
        """Passing user sets created_by on the created asset."""
        data = _build_xlsx([{
            "Asset Code": "IMP-USER", "Name": "With User",
            "Category Code": category.code,
        }])
        result = import_assets_from_excel(io.BytesIO(data), user=org_admin_user)
        assert result["success"] == 1
        a = Asset.objects.get(asset_code="IMP-USER")
        assert a.created_by_id == org_admin_user.id

    def test_multiple_rows_mixed_outcomes(self, asset_factory, category, tenant):
        """A single import mixing success, skip, and error rows."""
        asset_factory(asset_code="EXIST-1", category=category, serial_number="SN-X")
        data = _build_xlsx([
            {"Asset Code": "GOOD-1", "Name": "Good",
             "Category Code": category.code},
            {"Asset Code": "EXIST-1", "Name": "Skip",
             "Category Code": category.code, "Serial Number": "SN-X"},
            {"Asset Code": "BAD-1", "Name": "Bad", "Category Code": "NOPE"},
        ])
        result = import_assets_from_excel(io.BytesIO(data))
        assert result["success"] == 1
        assert result["skipped"] == 1
        assert len(result["errors"]) == 1

    def test_purchase_date_as_datetime_cell(self, category, tenant):
        """A real datetime cell value is coerced via .date() (parse_date branch)."""
        from datetime import datetime as _dt
        data = _build_xlsx([{
            "Asset Code": "IMP-DTCELL", "Name": "DT Cell",
            "Category Code": category.code,
            "Purchase Date": _dt(2023, 5, 1, 9, 30),
        }])
        result = import_assets_from_excel(io.BytesIO(data))
        assert result["success"] == 1
        a = Asset.objects.get(asset_code="IMP-DTCELL")
        assert a.purchase_date == date(2023, 5, 1)

    def test_invalid_warranty_date_format_error(self, category, tenant):
        """Malformed Warranty Expiry Date → its own parse-error branch."""
        data = _build_xlsx([{
            "Asset Code": "IMP-BADWAR", "Name": "Bad Warranty",
            "Category Code": category.code,
            "Warranty Expiry Date": "05/01/2023",
        }])
        result = import_assets_from_excel(io.BytesIO(data))
        assert result["success"] == 0
        assert any("Invalid date format" in e for e in result["errors"])

    def test_create_failure_recorded_as_row_error(self, category, tenant):
        """An unexpected DB error during create is caught and reported per-row."""
        from unittest.mock import patch as _patch
        data = _build_xlsx([{
            "Asset Code": "IMP-BOOM", "Name": "Boom",
            "Category Code": category.code,
        }])
        with _patch(
            "apps.assets.services.excel_service.Asset.objects.create",
            side_effect=Exception("db exploded"),
        ):
            result = import_assets_from_excel(io.BytesIO(data))
        assert result["success"] == 0
        assert any("Error parsing data" in e for e in result["errors"])


class TestExcelEndpoints:
    """API endpoints: /api/v1/assets/export_excel/ and import_excel/."""

    base_url = "/api/v1/assets/"

    def test_export_endpoint_returns_xlsx(self, hr_api_client, asset):
        response = hr_api_client.get(f"{self.base_url}export_excel/")
        assert response.status_code == status.HTTP_200_OK
        assert response["Content-Type"] == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert "attachment" in response["Content-Disposition"]
        ws = _load_ws(response.content)
        assert [c.value for c in ws[1]] == EXCEL_HEADERS

    def test_export_endpoint_requires_auth(self, api_client, tenant):
        response = api_client.get(f"{self.base_url}export_excel/")
        assert response.status_code in [
            status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN,
        ]

    def test_import_no_file_returns_400(self, hr_api_client, tenant):
        response = hr_api_client.post(f"{self.base_url}import_excel/", data={}, format="multipart")
        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert "No file provided" in response.data["error"]

    def test_import_wrong_extension_returns_400(self, hr_api_client, tenant):
        from django.core.files.uploadedfile import SimpleUploadedFile
        upload = SimpleUploadedFile(
            "assets.csv", b"some,data", content_type="text/csv"
        )
        response = hr_api_client.post(
            f"{self.base_url}import_excel/", data={"file": upload}, format="multipart"
        )
        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert "Invalid file format" in response.data["error"]

    def test_import_corrupt_xlsx_returns_400(self, hr_api_client, tenant):
        from django.core.files.uploadedfile import SimpleUploadedFile
        upload = SimpleUploadedFile(
            "bad.xlsx", b"not a real xlsx file",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response = hr_api_client.post(
            f"{self.base_url}import_excel/", data={"file": upload}, format="multipart"
        )
        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert "error" in response.data

    def test_import_zipfile_error_returns_friendly_message(self, hr_api_client, category, tenant):
        """A zipfile/workbook parse error maps to the 'corrupted file' 400 branch."""
        from unittest.mock import patch as _patch
        from django.core.files.uploadedfile import SimpleUploadedFile
        xlsx_bytes = _build_xlsx([{
            "Asset Code": "API-ZIP", "Name": "Zip",
            "Category Code": category.code,
        }])
        upload = SimpleUploadedFile(
            "assets.xlsx", xlsx_bytes,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        with _patch(
            "apps.assets.views.import_assets_from_excel",
            side_effect=Exception("File is not a valid zipfile"),
        ):
            response = hr_api_client.post(
                f"{self.base_url}import_excel/", data={"file": upload}, format="multipart"
            )
        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert "corrupted" in response.data["error"].lower()

    def test_import_valid_file_creates_assets(self, hr_api_client, category, tenant):
        from django.core.files.uploadedfile import SimpleUploadedFile
        xlsx_bytes = _build_xlsx([{
            "Asset Code": "API-IMP-1", "Name": "Via API",
            "Category Code": category.code,
        }])
        upload = SimpleUploadedFile(
            "assets.xlsx", xlsx_bytes,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response = hr_api_client.post(
            f"{self.base_url}import_excel/", data={"file": upload}, format="multipart"
        )
        assert response.status_code == status.HTTP_200_OK
        assert response.data["success"] == 1
        assert Asset.objects.filter(asset_code="API-IMP-1").exists()

    def test_import_forbidden_for_employee(self, employee_api_client, category, tenant):
        from django.core.files.uploadedfile import SimpleUploadedFile
        xlsx_bytes = _build_xlsx([{
            "Asset Code": "API-EMP", "Name": "Nope",
            "Category Code": category.code,
        }])
        upload = SimpleUploadedFile(
            "assets.xlsx", xlsx_bytes,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response = employee_api_client.post(
            f"{self.base_url}import_excel/", data={"file": upload}, format="multipart"
        )
        assert response.status_code == status.HTTP_403_FORBIDDEN


# ---------------------------------------------------------------------------
# Serializer branch coverage
# ---------------------------------------------------------------------------

class TestAssetSerializerBranches:
    """Cover AssetSerializer.to_representation and validate() branches."""

    url = "/api/v1/assets/"

    def test_representation_nests_category_and_null_metadata(self, hr_api_client, asset):
        response = hr_api_client.get(f"{self.url}{asset.id}/")
        assert response.status_code == status.HTTP_200_OK
        assert isinstance(response.data["category"], dict)
        assert response.data["category"]["id"] == str(asset.category_id)
        # metadata defaults to {} -> serialized as None
        assert response.data["metadata"] is None

    def test_representation_returns_metadata_when_present(self, hr_api_client, asset_factory, category):
        a = asset_factory(name="Meta Asset", category=category, metadata={"k": "v"})
        response = hr_api_client.get(f"{self.url}{a.id}/")
        assert response.data["metadata"] == {"k": "v"}

    def test_representation_nests_current_owner(self, hr_api_client, asset_factory, category, employee):
        a = asset_factory(name="Owned", category=category, current_owner=employee)
        response = hr_api_client.get(f"{self.url}{a.id}/")
        assert isinstance(response.data["current_owner"], dict)
        assert response.data["current_owner"]["id"] == str(employee.id)

    def test_representation_nests_current_allocation(
        self, hr_api_client, asset_factory, category, employee, allocation_factory
    ):
        a = asset_factory(name="Allocated", category=category)
        alloc = allocation_factory(asset=a, employee=employee)
        a.current_allocation = alloc
        a.save(update_fields=["current_allocation"])
        response = hr_api_client.get(f"{self.url}{a.id}/")
        assert isinstance(response.data["current_allocation"], dict)
        assert response.data["current_allocation"]["id"] == str(alloc.id)

    def test_warranty_before_purchase_rejected(self, hr_api_client, category):
        response = hr_api_client.post(self.url, data={
            "asset_code": "WAR-BAD", "name": "War",
            "category": str(category.id),
            "purchase_date": "2023-05-01",
            "warranty_expiry_date": "2022-05-01",
        })
        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert response.data.get("code") == 1
        assert "warranty_expiry_date" in response.data.get("message", "")

    def test_warranty_after_purchase_accepted(self, hr_api_client, category):
        response = hr_api_client.post(self.url, data={
            "asset_code": "WAR-OK", "name": "War",
            "category": str(category.id),
            "purchase_date": "2023-05-01",
            "warranty_expiry_date": "2025-05-01",
        })
        assert response.status_code == status.HTTP_201_CREATED

    def test_duplicate_serial_number_rejected(self, hr_api_client, asset_factory, category):
        asset_factory(name="First", category=category, serial_number="SER-UNIQUE")
        response = hr_api_client.post(self.url, data={
            "asset_code": "SER-DUP", "name": "Second",
            "category": str(category.id), "serial_number": "SER-UNIQUE",
        })
        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert response.data.get("code") == 2
        assert "message" in response.data

    def test_update_keeps_own_serial_number(self, hr_api_client, asset_factory, category):
        """Updating an asset without changing its serial must not self-conflict."""
        a = asset_factory(name="Keep", category=category, serial_number="SER-KEEP")
        response = hr_api_client.put(f"{self.url}{a.id}/", data={
            "asset_code": a.asset_code, "name": "Renamed",
            "category": str(category.id), "serial_number": "SER-KEEP",
        })
        assert response.status_code == status.HTTP_200_OK


class TestAssetCategorySerializerBranches:
    """Cover AssetCategorySerializer validation and representation branches."""

    url = "/api/v1/asset-categories/"

    def test_representation_nests_parent(self, hr_api_client, category_factory):
        parent = category_factory(name="Parent", code="PAR", category_type="HARDWARE")
        child = category_factory(name="Child", code="CHI", category_type="HARDWARE", parent=parent)
        response = hr_api_client.get(f"{self.url}{child.id}/")
        assert response.status_code == status.HTTP_200_OK
        assert isinstance(response.data["parent"], dict)
        assert response.data["parent"]["id"] == str(parent.id)

    def test_cannot_nest_under_subcategory(self, hr_api_client, category_factory):
        """Parent that is itself a sub-category is rejected (only 1 level)."""
        grandparent = category_factory(name="GP", code="GP", category_type="HARDWARE")
        parent = category_factory(name="P", code="P", category_type="HARDWARE", parent=grandparent)
        response = hr_api_client.post(self.url, data={
            "name": "Deep", "code": "DEEP", "category_type": "HARDWARE",
            "parent": str(parent.id),
        })
        assert response.status_code == status.HTTP_400_BAD_REQUEST

    def test_category_with_children_cannot_be_nested(self, hr_api_client, category_factory):
        """A category that already has children cannot become a sub-category."""
        parent = category_factory(name="HasKids", code="HK", category_type="HARDWARE")
        category_factory(name="Kid", code="KID", category_type="HARDWARE", parent=parent)
        target = category_factory(name="Other", code="OTH", category_type="HARDWARE")
        response = hr_api_client.put(f"{self.url}{parent.id}/", data={
            "name": "HasKids", "code": "HK", "category_type": "HARDWARE",
            "parent": str(target.id),
        })
        assert response.status_code == status.HTTP_400_BAD_REQUEST

    def test_subcategory_type_must_match_parent(self, hr_api_client, category_factory):
        parent = category_factory(name="HWParent", code="HWP", category_type="HARDWARE")
        response = hr_api_client.post(self.url, data={
            "name": "Mismatch", "code": "MIS", "category_type": "SOFTWARE",
            "parent": str(parent.id),
        })
        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert response.data.get("code") == 1
        assert "category_type" in response.data.get("message", "")


class TestAssetCategoryDestroy:
    """AssetCategoryViewSet.destroy — guards deleting a parent with active children."""

    url = "/api/v1/asset-categories/"

    def test_delete_category_without_children_succeeds(self, hr_api_client, category_factory):
        """A leaf category deletes cleanly."""
        cat = category_factory(name="Solo", code="SOLO", category_type="HARDWARE")
        response = hr_api_client.delete(f"{self.url}{cat.id}/")
        assert response.status_code in [
            status.HTTP_200_OK, status.HTTP_204_NO_CONTENT,
        ]

    def test_delete_category_with_active_children_rejected(self, hr_api_client, category_factory):
        """A parent with an active child → 400 with an explanatory message."""
        parent = category_factory(name="Parent", code="PARENT", category_type="HARDWARE")
        category_factory(
            name="Kid", code="KIDX", category_type="HARDWARE", parent=parent
        )
        response = hr_api_client.delete(f"{self.url}{parent.id}/")
        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert "child categories" in response.data["message"].lower()

    def test_delete_nonexistent_category_returns_404(self, hr_api_client, tenant):
        """A missing id is caught and delegated to the default 404 flow."""
        import uuid as _uuid
        response = hr_api_client.delete(f"{self.url}{_uuid.uuid4()}/")
        assert response.status_code == status.HTTP_404_NOT_FOUND
